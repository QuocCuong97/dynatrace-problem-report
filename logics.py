import time
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from common import load_from_json, convert_to_date_object


def main(api_from ,api_to, report_from, report_to, sort):
    settings = load_from_json('settings.json')

    def get_list_problems():
        output = requests.get(settings['dynatrace']['url'] + '/api/v2/problems?from={}&to={}&sort={}&pageSize=500'.format(api_from, api_to, sort), \
                            headers={"Authorization": "Api-Token {}".format(settings['dynatrace']['api-token'])}).json()
        problems = output['problems']
        return problems

    def export_report(problems):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Problem'
        wbd = Side(border_style='thick', color="00FFFFFF")
        bd = Side(border_style='thin', color="000000")
        
        # Header
        ws.merge_cells('A1:A3')
        ws['A1'].border = Border(left=wbd, top=wbd, right=wbd)
        ws['A2'].border = Border(left=wbd, top=wbd, right=wbd)
        ws['A3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        
        ws.merge_cells('B1:C3')
        img = Image('report.png')
        ws.add_image(img, 'B1')
        ws['B1'].border = Border(left=wbd, top=wbd, right=wbd)
        ws['C2'].border = Border(left=wbd, top=wbd, right=wbd)
        ws['B3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['C3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)

        ws.merge_cells('D1:L2')
        ws['D1'].font = Font(bold=True, size=13)
        ws['D1'].alignment = Alignment(horizontal='center', vertical="center")
        ws['D1'] = settings['output_file']['title']
        ws['D1'].border = Border(left=wbd, top=wbd, right=wbd)
        redFill = PatternFill(start_color='0099CCFF',
                   end_color='0099CCFF',
                   fill_type='solid')
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[5].height = 20
        ws['A5'].fill = redFill
        ws['B5'].fill = redFill
        ws['C5'].fill = redFill
        ws['D5'].fill = redFill
        ws['E5'].fill = redFill
        ws['F5'].fill = redFill
        ws['G5'].fill = redFill
        ws['H5'].fill = redFill
        ws['I5'].fill = redFill
        ws['J5'].fill = redFill
        ws['K5'].fill = redFill
        ws['L5'].fill = redFill
        
        
        ws.merge_cells('D3:L3')
        ws['D3'].font = Font(bold=True, italic=True)
        ws['D3'].alignment = Alignment(horizontal='center', vertical="center")
        ws['D3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['E3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['F3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['G3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['H3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['I3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['J3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['K3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['L3'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['L1'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['L2'].border = Border(left=wbd, top=wbd, right=wbd, bottom=wbd)
        ws['L4'].border = Border(left=wbd, top=wbd, right=wbd)
        ws['D3'] = "Total: {} problems\
                             (from {} to {})".format(len(problems), report_from, report_to)

        ws.merge_cells('A4:L4')

        # Body
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 27
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 13
        ws.column_dimensions['L'].width = 10
        columns = ['STT', 'Problem ID', 'Title', 'Severity Level', 'Impact Level', 'Affected', 'Impacted', 'Root cause', 'Start time', 'End time', 'Duration', 'Status']
        
        row_num = 5
        ordinary = 0
        
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical="center")
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        for problem in problems:
            row_num += 1
            ordinary += 1
            display_id = problem['displayId']
            title = problem['title']
            severity_level = problem['severityLevel']
            impact_level = problem['impactLevel']
            status = problem['status']

            start_time = datetime.fromtimestamp(int(problem['startTime']) / 1000).strftime("%B %d, %Y %H:%M")
            end_time = datetime.fromtimestamp(int(problem['endTime']) / 1000).strftime("%B %d, %Y %H:%M")

            duration_raw = time.gmtime((int(problem['endTime']) - int(problem['startTime'])) / 1000)
            if duration_raw.tm_hour == 0:
                duration = time.strftime("%M min", duration_raw)
            if duration_raw.tm_hour != 0:
                duration = time.strftime("%H hour %M min", duration_raw)
            if duration[0] == "0":
                duration = duration[1:]
            
            if problem['rootCauseEntity'] is None:
                root_cause = ""
            else:
                root_cause = problem['rootCauseEntity']['name']
            
            affected = len(problem['impactedEntities'])
            impacted = ""
            for entity in problem['impactedEntities']:
                impacted += "{}, ".format(entity['name'])
            impacted = impacted[:-2]

            row = [ordinary, display_id, title, severity_level, impact_level, affected, impacted, root_cause, start_time, end_time, duration, status]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center", wrapText=True)
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        
        file_ouput = settings['output_file']['output_folder'] + "/" + \
                     settings['output_file']['file_name_prefix'] +  \
                         "_" + convert_to_date_object(report_from) + "_" + \
                             convert_to_date_object(report_to) + ".xlsx"
        wb.save(filename=file_ouput)
        return file_ouput

    list_problems = get_list_problems()
    if len(list_problems) != 0:
        output = export_report(list_problems)
        return output
    else:
        raise Exception("No problems found!")
